"""
Excel Şablon Oluşturucu v2
Veri girişi için boş Excel şablonu oluşturur.
Güncellemeler:
- Ana Dağıtım Pano bilgileri ayrı sayfa
- Tüm kontrol kriterleri eklendi
- Dropdown menüler (Açma eğrisi, sigorta akımları vb.)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os


def create_excel_template(output_path: str):
    """Veri girişi için Excel şablonu oluşturur."""

    wb = Workbook()

    # Stiller
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === 1. FirmaBilgileri Sayfası ===
    ws_firma = wb.active
    ws_firma.title = "FirmaBilgileri"

    firma_headers = ["Alan Adı", "Değer"]
    firma_fields = [
        ("Firma Adi", ""),
        ("Adres", ""),
        ("Rapor Numarasi", ""),
        ("Rapor Tarihi", ""),
        ("ISG Katip ID", ""),
        ("Kontrol Baslangic", ""),
        ("Kontrol Bitis", ""),
        ("SGK Sicil", ""),
        ("Sonraki Kontrol", ""),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_firma.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(firma_fields, 2):
        ws_firma.cell(row=row, column=1, value=field).border = thin_border
        ws_firma.cell(row=row, column=2, value=value).border = thin_border

    ws_firma.column_dimensions['A'].width = 25
    ws_firma.column_dimensions['B'].width = 40

    # === 2. AnaDagitimPano Sayfası (YENİ - 2.1 DETAY BİLGİLER) ===
    ws_ana_pano = wb.create_sheet("AnaDagitimPano")

    ana_pano_fields = [
        ("Enerji Saglayan Kurulus", ""),
        ("Sebeke Tipi", "TT / IT / TN / TN-CS / TN-C / TN-S"),
        ("Sebeke Gerilimi", "380V"),
        ("Tesise Ait Proje", "Var / Yok"),
        ("Tek Hat Semasi", "Var / Yok"),
        ("Kontrol Nedeni", "Periyodik Kontrol / İlk Kontrol"),
        ("Topraklayici Tipi", "Ring / Yüzeysel / Temel / Derin / Belirlenemedi"),
        ("Yapi Cinsi", "Ev / Ticari / Endüstri / Diğer"),
        ("Temel Topraklama Direnci (Ohm)", ""),
        ("Ilave Topraklama Elektrotu Detaylari", ""),
        ("Sistem Topraklama Iletkeni Kesiti (mm2)", ""),
        ("Ana Espotansiyel Iletkeni Kesiti (mm2)", ""),
        ("Faz Iletkenleri Sayisi ve Tipi", "AA / 1 faz, 2 tel / 1 faz, 3 tel / 3 faz, 3 tel / 3 faz, 4 tel"),
        ("Besleme Kaynagi Nominal Gerilimi", "0,4/0,23 kV"),
        ("Dis Cevrim Empedansi Z_E (Ohm)", ""),  # I_f = 230/Z_E otomatik hesaplanacak
        ("Ana Kesici Tipi", "B / C / D"),
        ("Ana Kesici Nominal Akimi", ""),
        ("Ana RCD Tipi", "KAKR / TOROİD"),
        ("Ana RCD Dayanim Akimi (mA)", ""),
        ("Ana RCD Test Akimi (mA)", ""),
        ("Ana RCD Acma Suresi (ms)", ""),
        ("Tesisatta Kapsamli Degisiklik Var Mi (>%20)", "Var / Yok"),
        ("Asiri Gerilim Koruma Cihazi (DKD/SPD)", "Evet / Hayır"),
        ("Dogrudan Dokunmaya Karsi Koruma Onlemi", ""),
        ("Bir Onceki Periyodik Kontrol Etiketi", "Var / Yok"),
        # Fonksiyon Testleri Placeholder Alanları
        ("Pano Adi (PANO_adi1)", ""),
        ("Faz-Toprak Cevrim Empedansi Z_x (Ohm)", ""),
        ("Faz-Notr Cevrim Empedansi Z_ln (Ohm)", ""),
        ("Gerilim F-F (V)", ""),
        ("Gerilim L-N (V)", ""),
        ("Gerilim N-PE (V)", ""),
        ("Parafudr Tipi", ""),
        ("Parafudr Imax (kA)", ""),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_ana_pano.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(ana_pano_fields, 2):
        ws_ana_pano.cell(row=row, column=1, value=field).border = thin_border
        ws_ana_pano.cell(row=row, column=2, value=value).border = thin_border

    ws_ana_pano.column_dimensions['A'].width = 45
    ws_ana_pano.column_dimensions['B'].width = 50

    # === 3. CihazBilgileri Sayfası ===
    ws_cihaz = wb.create_sheet("CihazBilgileri")

    cihaz_fields = [
        ("--- TERMAL KAMERA ---", ""),
        ("Cihaz Adi", ""),
        ("Kalibrasyon Tarihi", ""),
        ("Gecerlilik Tarihi", ""),
        ("Seri No", ""),
        ("Kalibrasyon No", ""),
        ("--- ÖLÇÜM ALETİ ---", ""),
        ("Cihaz Adi", ""),
        ("Kalibrasyon Tarihi", ""),
        ("Gecerlilik Tarihi", ""),
        ("Seri No", ""),
        ("Kalibrasyon No", ""),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_cihaz.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(cihaz_fields, 2):
        ws_cihaz.cell(row=row, column=1, value=field).border = thin_border
        ws_cihaz.cell(row=row, column=2, value=value).border = thin_border

    ws_cihaz.column_dimensions['A'].width = 25
    ws_cihaz.column_dimensions['B'].width = 35

    # === 4. GozleKontrol Sayfası (TÜM MADDELER) ===
    ws_kontrol = wb.create_sheet("GozleKontrol")

    kontrol_headers = ["Kontrol Kriteri", "Değerlendirme", "Kusur Derecesi"]

    # Tablo 2'deki TÜM maddeler
    kontrol_fields = [
        ("PANO_ADI", "", ""),
        ("--- PANO VE DİĞER DONANIMLARA GİRİŞİN UYGUNLUĞU ---", "", ""),
        ("Kablo Sebeke Tarafi", "Uygun", ""),
        ("Kablo Donanim Tarafi", "Uygun", ""),
        ("Pano Sabitlenmesi (Depreme Dayaniklilik)", "Uygun", ""),
        ("Dis Darbelere Karsi Koruma Onlemi", "Uygun", ""),
        ("Elektrik Panosu Etrafinda Yabanci Malzemeler", "Uygun", ""),
        ("Zemin Izolasyonu", "Uygun", ""),
        ("--- TOPRAKLANMIŞ POTANSİYEL DENGELEME VE BESLEMENİN OTOMATİK KESİLMESİ ---", "", ""),
        ("Topraklama Iletkeni", "Uygun", ""),
        ("Ana Potansiyel Dengeleme Iletkeni", "Uygun", ""),
        ("Ek Potansiyel Dengeleme Iletkeni (Tamamlayici)", "Uygun", ""),
        ("Pano Kapak Baglantisi Kontrolu 6 mm2", "Uygun", ""),
        ("--- KARŞILIKLI ZARARLI ETKİLERİN ÖNLENMESİ ---", "", ""),
        ("Elektriksel Olmayan Tesislere Yaklasma ve Diger Etkilerin Kontrolu", "Uygun", ""),
        ("Bant I ve Bant II Ayrilmasi, Bant II Yalitimi", "Uygun", ""),
        ("Guvenlik Devre Ayrilmasi", "Uygun", ""),
        ("Pano Ic Kapak, Faza Erisim Engeli veya Pleksi Koruma", "Uygun", ""),
        ("--- TANIMLAMA ---", "", ""),
        ("Semalar, Talimatlar, Devre Cizimleri ve Kisa Bilgiler", "Uygun", ""),
        ("Koruma Cihaz ve Terminal Etiket", "Uygun", ""),
        ("Tehlike Isaretleri ve Diger Uyari Isaretleri", "Uygun", ""),
        ("--- KABLO ve İLETKENLER ---", "", ""),
        ("Kablo Yollarinin Uygunlugu ve Mekanik Koruma", "Uygun", ""),
        ("Kablo Renk Kodlari (Notr: Mavi, Toprak: Sari/Yesil)", "Uygun", ""),
        ("Tesisat Yontemi", "Uygun", ""),
        ("Yangin Engeli, Uygun Kilitleme ve Sicaklik Etkisine Karsi Koruma", "Uygun", ""),
        ("--- TERMAL KAMERA ---", "", ""),
        ("Fotograf Tarihi", "", ""),
        ("Fotograf No", "", ""),
        ("Kontak Gevsekligi Isinmasi", "Uygun", ""),
        ("Asiri Yuk Isinmasi (PVC Kablolar Icin >70 derece)", "Uygun", ""),
        ("--- GENEL DEĞERLENDİRMELER ---", "", ""),
        ("Ekipman Yakininda Elektriksel Ekipman Yangin Sondurme Tertibati", "Uygun", ""),
        ("Ekipman Temizlik/Bakim Durumu", "Uygun", ""),
        ("Pano Ici ve Baglantilarinin Korozyon Kontrolu", "Uygun", ""),
        ("Ekipman Ici veya Yakininda Acil Durum Aydinlatma Tertibati", "Uygun", ""),
    ]

    for col, header in enumerate(kontrol_headers, 1):
        cell = ws_kontrol.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value, kusur) in enumerate(kontrol_fields, 2):
        cell1 = ws_kontrol.cell(row=row, column=1, value=field)
        cell1.border = thin_border
        if field.startswith("---"):
            cell1.fill = section_fill
            cell1.font = Font(bold=True)

        ws_kontrol.cell(row=row, column=2, value=value).border = thin_border
        ws_kontrol.cell(row=row, column=3, value=kusur).border = thin_border

    # Dropdown for Değerlendirme
    dv_degerlendirme = DataValidation(
        type="list",
        formula1='"Uygun,Uygun Değil,Uygulanamaz"',
        allow_blank=True
    )
    dv_degerlendirme.error = "Listeden seçim yapın"
    dv_degerlendirme.errorTitle = "Geçersiz değer"
    ws_kontrol.add_data_validation(dv_degerlendirme)
    dv_degerlendirme.add(f"B2:B{len(kontrol_fields)+1}")

    # Dropdown for Kusur Derecesi
    dv_kusur = DataValidation(
        type="list",
        formula1='"*,**,"',
        allow_blank=True
    )
    ws_kontrol.add_data_validation(dv_kusur)
    dv_kusur.add(f"C2:C{len(kontrol_fields)+1}")

    ws_kontrol.column_dimensions['A'].width = 60
    ws_kontrol.column_dimensions['B'].width = 20
    ws_kontrol.column_dimensions['C'].width = 15

    # === 5. FonksiyonTestleri Sayfası (DROPDOWN MENÜLER) ===
    ws_fonksiyon = wb.create_sheet("FonksiyonTestleri")

    fonksiyon_headers = [
        "No", "Linye Adi", "Acma Egrisi", "Kutup Sayisi", "In (A)",
        "Icu (kA)", "Faz Kesiti (mm2)", "N Kesiti (mm2)", "PE Kesiti (mm2)",
        "Ib (A)", "Iz (A)", "RCD mA", "RCD ms", "Sonuc"
    ]

    for col, header in enumerate(fonksiyon_headers, 1):
        cell = ws_fonksiyon.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # 100 boş satır ekle (dinamik yapı)
    for row in range(2, 102):
        ws_fonksiyon.cell(row=row, column=1, value=row-1).border = thin_border
        for col in range(2, len(fonksiyon_headers) + 1):
            ws_fonksiyon.cell(row=row, column=col, value="").border = thin_border

    # Dropdown: Açma Eğrisi Tipi
    dv_egri = DataValidation(
        type="list",
        formula1='"B,C,D,K,Z,KAKR"',
        allow_blank=True
    )
    dv_egri.error = "B, C, D, K veya Z seçin"
    ws_fonksiyon.add_data_validation(dv_egri)
    dv_egri.add("C2:C101")

    # Dropdown: Kutup Sayısı
    dv_kutup = DataValidation(
        type="list",
        formula1='"1,2,3,4"',
        allow_blank=True
    )
    ws_fonksiyon.add_data_validation(dv_kutup)
    dv_kutup.add("D2:D101")

    # Dropdown: Nominal Akım (In) - Standart sigorta değerleri
    dv_in = DataValidation(
        type="list",
        formula1='"6,10,13,16,20,25,32,40,50,63,80,100,125,160,200,250,315,400,500,630"',
        allow_blank=True
    )
    ws_fonksiyon.add_data_validation(dv_in)
    dv_in.add("E2:E101")

    # Dropdown: Kısa devre kesme kapasitesi (Icu)
    dv_icu = DataValidation(
        type="list",
        formula1='"4.5,6,10,15,25,36,50"',
        allow_blank=True
    )
    ws_fonksiyon.add_data_validation(dv_icu)
    dv_icu.add("F2:F101")

    # Dropdown: Kesit değerleri
    dv_kesit = DataValidation(
        type="list",
        formula1='"1.5,2.5,4,6,10,16,25,35,50,70,95,120,150,185,240"',
        allow_blank=True
    )
    ws_fonksiyon.add_data_validation(dv_kesit)
    dv_kesit.add("G2:G101")  # Faz kesiti
    dv_kesit2 = DataValidation(type="list", formula1='"1.5,2.5,4,6,10,16,25,35,50,70,95,120,150,185,240"', allow_blank=True)
    ws_fonksiyon.add_data_validation(dv_kesit2)
    dv_kesit2.add("H2:H101")  # N kesiti
    dv_kesit3 = DataValidation(type="list", formula1='"1.5,2.5,4,6,10,16,25,35,50,70,95,120,150,185,240"', allow_blank=True)
    ws_fonksiyon.add_data_validation(dv_kesit3)
    dv_kesit3.add("I2:I101")  # PE kesiti

    # Dropdown: RCD mA değerleri
    dv_rcd_ma = DataValidation(
        type="list",
        formula1='"10,30,100,300,500"',
        allow_blank=True
    )
    ws_fonksiyon.add_data_validation(dv_rcd_ma)
    dv_rcd_ma.add("L2:L101")

    # Dropdown: Sonuç
    dv_sonuc = DataValidation(
        type="list",
        formula1='"Uygun,Uygun Değil"',
        allow_blank=True
    )
    ws_fonksiyon.add_data_validation(dv_sonuc)
    dv_sonuc.add("N2:N101")

    # Sütun genişlikleri
    for col in range(1, len(fonksiyon_headers) + 1):
        ws_fonksiyon.column_dimensions[get_column_letter(col)].width = 12
    ws_fonksiyon.column_dimensions['B'].width = 30  # Linye Adı

    # === 6. TermalGoruntuler Sayfası ===
    ws_termal = wb.create_sheet("TermalGoruntuler")

    termal_headers = ["Pano Adı", "Fluke DOCX Dosya Yolu"]

    for col, header in enumerate(termal_headers, 1):
        cell = ws_termal.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row in range(2, 7):
        ws_termal.cell(row=row, column=1, value="").border = thin_border
        ws_termal.cell(row=row, column=2, value="").border = thin_border

    ws_termal.column_dimensions['A'].width = 25
    ws_termal.column_dimensions['B'].width = 60

    # === 7. Sonuc Sayfası ===
    ws_sonuc = wb.create_sheet("Sonuc")

    sonuc_fields = [
        ("Ek Notlar", ""),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_sonuc.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(sonuc_fields, 2):
        ws_sonuc.cell(row=row, column=1, value=field).border = thin_border
        ws_sonuc.cell(row=row, column=2, value=value).border = thin_border

    # Not
    ws_sonuc.cell(row=4, column=1, value="NOT: Kusur açıklamaları GozleKontrol sayfasındaki 'Uygun Değil' maddelerinden otomatik oluşturulur.")

    ws_sonuc.column_dimensions['A'].width = 25
    ws_sonuc.column_dimensions['B'].width = 80

    # Kaydet
    wb.save(output_path)
    print(f"Excel sablonu olusturuldu: {output_path}")
    return output_path


def create_sample_data(output_path: str):
    """Örnek verilerle dolu Excel dosyası oluşturur.

    NOT: veri_sablonu.xlsx ile AYNI yapıda olmalı, sadece örnek veriler dolu.
    """

    wb = Workbook()

    # Stiller
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === 1. FirmaBilgileri Sayfası ===
    ws_firma = wb.active
    ws_firma.title = "FirmaBilgileri"

    firma_headers = ["Alan Adı", "Değer"]
    firma_fields = [
        ("Firma Adi", "MİKROPOR SANAYİ A.Ş."),
        ("Adres", "Organize Sanayi Bölgesi 7. Cadde No:15"),
        ("Rapor Numarasi", "2025-3931-2"),
        ("Rapor Tarihi", "12.12.2025"),
        ("ISG Katip ID", "ISG-2025-001234"),
        ("Kontrol Baslangic", "12.12.2025 09:00"),
        ("Kontrol Bitis", "12.12.2025 17:00"),
        ("SGK Sicil", "1234567890"),
        ("Sonraki Kontrol", "12.12.2026"),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_firma.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(firma_fields, 2):
        ws_firma.cell(row=row, column=1, value=field).border = thin_border
        ws_firma.cell(row=row, column=2, value=value).border = thin_border

    ws_firma.column_dimensions['A'].width = 25
    ws_firma.column_dimensions['B'].width = 40

    # === 2. AnaDagitimPano Sayfası ===
    ws_ana_pano = wb.create_sheet("AnaDagitimPano")

    ana_pano_fields = [
        ("Enerji Saglayan Kurulus", "TEDAŞ"),
        ("Sebeke Tipi", "TT"),
        ("Sebeke Gerilimi", "380V"),
        ("Tesise Ait Proje", "Var"),
        ("Tek Hat Semasi", "Var"),
        ("Kontrol Nedeni", "Periyodik Kontrol"),
        ("Topraklayici Tipi", "Temel"),
        ("Yapi Cinsi", "Endüstri"),
        ("Temel Topraklama Direnci (Ohm)", "2.5"),
        ("Ilave Topraklama Elektrotu Detaylari", "-"),
        ("Sistem Topraklama Iletkeni Kesiti (mm2)", "16"),
        ("Ana Espotansiyel Iletkeni Kesiti (mm2)", "10"),
        ("Faz Iletkenleri Sayisi ve Tipi", "3 faz, 4 tel"),
        ("Besleme Kaynagi Nominal Gerilimi", "0,4/0,23 kV"),
        ("Dis Cevrim Empedansi Z_E (Ohm)", "0.35"),  # I_f = 230/0.35 = 657 A
        ("Ana Kesici Tipi", "C"),
        ("Ana Kesici Nominal Akimi", "250A"),
        ("Ana RCD Tipi", "TOROİD"),
        ("Ana RCD Dayanim Akimi (mA)", "300"),
        ("Ana RCD Test Akimi (mA)", "150"),
        ("Ana RCD Acma Suresi (ms)", "20"),
        ("Tesisatta Kapsamli Degisiklik Var Mi (>%20)", "Yok"),
        ("Asiri Gerilim Koruma Cihazi (DKD/SPD)", "Evet"),
        ("Dogrudan Dokunmaya Karsi Koruma Onlemi", "İç kapak"),
        ("Bir Onceki Periyodik Kontrol Etiketi", "Var"),
        # Fonksiyon Testleri Placeholder Alanları
        ("Pano Adi (PANO_adi1)", "ANA DAĞITIM PANOSU"),
        ("Faz-Toprak Cevrim Empedansi Z_x (Ohm)", "0.42"),
        ("Faz-Notr Cevrim Empedansi Z_ln (Ohm)", "0.28"),
        ("Gerilim F-F (V)", "398"),
        ("Gerilim L-N (V)", "230"),
        ("Gerilim N-PE (V)", "0.5"),
        ("Parafudr Tipi", "Tip 2"),
        ("Parafudr Imax (kA)", "40"),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_ana_pano.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(ana_pano_fields, 2):
        ws_ana_pano.cell(row=row, column=1, value=field).border = thin_border
        ws_ana_pano.cell(row=row, column=2, value=value).border = thin_border

    ws_ana_pano.column_dimensions['A'].width = 45
    ws_ana_pano.column_dimensions['B'].width = 50

    # === 3. CihazBilgileri Sayfası ===
    ws_cihaz = wb.create_sheet("CihazBilgileri")

    cihaz_fields = [
        ("--- TERMAL KAMERA ---", ""),
        ("Cihaz Adi", "FLUKE TC01A"),
        ("Kalibrasyon Tarihi", "04.09.2025"),
        ("Gecerlilik Tarihi", "04.09.2026"),
        ("Seri No", "69401214"),
        ("Kalibrasyon No", "K-AF16-6C08"),
        ("--- ÖLÇÜM ALETİ ---", ""),
        ("Cihaz Adi", "FLUKE 1663 SCH"),
        ("Kalibrasyon Tarihi", "02.10.2025"),
        ("Gecerlilik Tarihi", "02.10.2026"),
        ("Seri No", "5436127"),
        ("Kalibrasyon No", "0039K-1025-00114"),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_cihaz.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(cihaz_fields, 2):
        ws_cihaz.cell(row=row, column=1, value=field).border = thin_border
        ws_cihaz.cell(row=row, column=2, value=value).border = thin_border

    ws_cihaz.column_dimensions['A'].width = 25
    ws_cihaz.column_dimensions['B'].width = 35

    # === 4. GozleKontrol Sayfası (TÜM MADDELER - veri_sablonu ile AYNI) ===
    ws_kontrol = wb.create_sheet("GozleKontrol")

    kontrol_headers = ["Kontrol Kriteri", "Değerlendirme", "Kusur Derecesi"]

    # TÜM maddeler - örnek değerlerle
    kontrol_fields = [
        ("PANO_ADI", "A06 TEZGAH", ""),
        ("--- PANO VE DİĞER DONANIMLARA GİRİŞİN UYGUNLUĞU ---", "", ""),
        ("Kablo Sebeke Tarafi", "Uygun", ""),
        ("Kablo Donanim Tarafi", "Uygun", ""),
        ("Pano Sabitlenmesi (Depreme Dayaniklilik)", "Uygun", ""),
        ("Dis Darbelere Karsi Koruma Onlemi", "Uygun", ""),
        ("Elektrik Panosu Etrafinda Yabanci Malzemeler", "Uygun Değil", "*"),
        ("Zemin Izolasyonu", "Uygun", ""),
        ("--- TOPRAKLANMIŞ POTANSİYEL DENGELEME VE BESLEMENİN OTOMATİK KESİLMESİ ---", "", ""),
        ("Topraklama Iletkeni", "Uygun", ""),
        ("Ana Potansiyel Dengeleme Iletkeni", "Uygun", ""),
        ("Ek Potansiyel Dengeleme Iletkeni (Tamamlayici)", "Uygun", ""),
        ("Pano Kapak Baglantisi Kontrolu 6 mm2", "Uygun", ""),
        ("--- KARŞILIKLI ZARARLI ETKİLERİN ÖNLENMESİ ---", "", ""),
        ("Elektriksel Olmayan Tesislere Yaklasma ve Diger Etkilerin Kontrolu", "Uygun", ""),
        ("Bant I ve Bant II Ayrilmasi, Bant II Yalitimi", "Uygun", ""),
        ("Guvenlik Devre Ayrilmasi", "Uygun", ""),
        ("Pano Ic Kapak, Faza Erisim Engeli veya Pleksi Koruma", "Uygun", ""),
        ("--- TANIMLAMA ---", "", ""),
        ("Semalar, Talimatlar, Devre Cizimleri ve Kisa Bilgiler", "Uygun", ""),
        ("Koruma Cihaz ve Terminal Etiket", "Uygun", ""),
        ("Tehlike Isaretleri ve Diger Uyari Isaretleri", "Uygun", ""),
        ("--- KABLO ve İLETKENLER ---", "", ""),
        ("Kablo Yollarinin Uygunlugu ve Mekanik Koruma", "Uygun", ""),
        ("Kablo Renk Kodlari (Notr: Mavi, Toprak: Sari/Yesil)", "Uygun", ""),
        ("Tesisat Yontemi", "Uygun", ""),
        ("Yangin Engeli, Uygun Kilitleme ve Sicaklik Etkisine Karsi Koruma", "Uygun", ""),
        ("--- TERMAL KAMERA ---", "", ""),
        ("Fotograf Tarihi", "12.12.2025", ""),
        ("Fotograf No", "16933E143", ""),
        ("Kontak Gevsekligi Isinmasi", "Uygun", ""),
        ("Asiri Yuk Isinmasi (PVC Kablolar Icin >70 derece)", "Uygun", ""),
        ("--- GENEL DEĞERLENDİRMELER ---", "", ""),
        ("Ekipman Yakininda Elektriksel Ekipman Yangin Sondurme Tertibati", "Uygun", ""),
        ("Ekipman Temizlik/Bakim Durumu", "Uygun", ""),
        ("Pano Ici ve Baglantilarinin Korozyon Kontrolu", "Uygun", ""),
        ("Ekipman Ici veya Yakininda Acil Durum Aydinlatma Tertibati", "Uygun", ""),
    ]

    for col, header in enumerate(kontrol_headers, 1):
        cell = ws_kontrol.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value, kusur) in enumerate(kontrol_fields, 2):
        cell1 = ws_kontrol.cell(row=row, column=1, value=field)
        cell1.border = thin_border
        if field.startswith("---"):
            cell1.fill = section_fill
            cell1.font = Font(bold=True)

        ws_kontrol.cell(row=row, column=2, value=value).border = thin_border
        ws_kontrol.cell(row=row, column=3, value=kusur).border = thin_border

    ws_kontrol.column_dimensions['A'].width = 60
    ws_kontrol.column_dimensions['B'].width = 20
    ws_kontrol.column_dimensions['C'].width = 15

    # === 5. FonksiyonTestleri Sayfası ===
    ws_fonksiyon = wb.create_sheet("FonksiyonTestleri")

    fonksiyon_headers = [
        "No", "Linye Adi", "Acma Egrisi", "Kutup Sayisi", "In (A)",
        "Icu (kA)", "Faz Kesiti (mm2)", "N Kesiti (mm2)", "PE Kesiti (mm2)",
        "Ib (A)", "Iz (A)", "RCD mA", "RCD ms", "Sonuc"
    ]

    for col, header in enumerate(fonksiyon_headers, 1):
        cell = ws_fonksiyon.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Örnek fonksiyon test verileri
    sample_rows = [
        [1, "Ana Kesici", "C", 3, 63, 10, 16, 16, 6, 45, "", 300, 25, "Uygun"],
        [2, "Aydınlatma", "B", 1, 16, 6, 2.5, 2.5, 2.5, 8, "", 30, 18, "Uygun"],
        [3, "Priz Devresi", "C", 1, 16, 6, 2.5, 2.5, 2.5, 12, "", 30, 22, "Uygun"],
        [4, "Klima", "C", 1, 20, 6, 4, 4, 4, 15, "", 30, 20, "Uygun"],
    ]

    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_fonksiyon.cell(row=row_idx, column=col_idx, value=value).border = thin_border

    # 22 satıra kadar boş satır ekle
    for row in range(len(sample_rows) + 2, 24):
        ws_fonksiyon.cell(row=row, column=1, value=row-1).border = thin_border
        for col in range(2, len(fonksiyon_headers) + 1):
            ws_fonksiyon.cell(row=row, column=col, value="").border = thin_border

    # Sütun genişlikleri
    for col in range(1, len(fonksiyon_headers) + 1):
        ws_fonksiyon.column_dimensions[get_column_letter(col)].width = 12
    ws_fonksiyon.column_dimensions['B'].width = 30

    # === 6. TermalGoruntuler Sayfası ===
    ws_termal = wb.create_sheet("TermalGoruntuler")

    termal_headers = ["Pano Adı", "Fluke DOCX Dosya Yolu"]

    for col, header in enumerate(termal_headers, 1):
        cell = ws_termal.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Örnek termal görüntü verisi
    ws_termal.cell(row=2, column=1, value="A06 TEZGAH").border = thin_border
    ws_termal.cell(row=2, column=2, value=r"c:\Users\cmshe\OneDrive\Masaüstü\BUILDV\FLUKE-16933E143.docx").border = thin_border

    for row in range(3, 7):
        ws_termal.cell(row=row, column=1, value="").border = thin_border
        ws_termal.cell(row=row, column=2, value="").border = thin_border

    ws_termal.column_dimensions['A'].width = 25
    ws_termal.column_dimensions['B'].width = 60

    # === 7. Sonuc Sayfası ===
    ws_sonuc = wb.create_sheet("Sonuc")

    sonuc_fields = [
        ("Ek Notlar", ""),
    ]

    for col, header in enumerate(firma_headers, 1):
        cell = ws_sonuc.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row, (field, value) in enumerate(sonuc_fields, 2):
        ws_sonuc.cell(row=row, column=1, value=field).border = thin_border
        ws_sonuc.cell(row=row, column=2, value=value).border = thin_border

    # Not
    ws_sonuc.cell(row=4, column=1, value="NOT: Kusur açıklamaları GozleKontrol sayfasındaki 'Uygun Değil' maddelerinden otomatik oluşturulur.")

    ws_sonuc.column_dimensions['A'].width = 25
    ws_sonuc.column_dimensions['B'].width = 80

    # Kaydet
    wb.save(output_path)
    print(f"Ornek veri dosyasi olusturuldu: {output_path}")
    return output_path


if __name__ == "__main__":
    base_dir = r"c:\Users\cmshe\OneDrive\Masaüstü\BUILDV\rapor_sistemi"

    # Boş şablon oluştur
    template_path = os.path.join(base_dir, "veri_sablonu.xlsx")
    create_excel_template(template_path)

    # Örnek veri dosyası oluştur
    sample_path = os.path.join(base_dir, "ornek_veri.xlsx")
    create_sample_data(sample_path)
