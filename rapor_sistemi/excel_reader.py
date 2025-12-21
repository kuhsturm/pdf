"""
Excel Okuyucu Modülü v2
Rapor verilerini Excel dosyasından okur.
Güncellemeler:
- AnaDagitimPano sayfası desteği
- Kusur derecesi desteği
- Iz otomatik hesaplama
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import os
import rapor_sistemi.constants as const

# Iz (Akım taşıma kapasitesi) tablosu - 2. Grup
IZ_TABLE_GROUP2 = const.IZ_TABLE


def parse_kesit_with_multiplier(kesit_val) -> float:
    """'2x16' gibi değerleri çarpanla parçalayıp efektif kesiti döndürür."""
    # Use the function from constants for consistency
    return const.parse_kesit_value(kesit_val)


def get_iz_from_kesit(kesit) -> int:
    """Faz kesitinden Iz (Akım taşıma kapasitesi) değerini hesaplar."""
    iz_str = const.calculate_iz(kesit)
    try:
        return int(iz_str)
    except ValueError:
        return 0


class ExcelReader:
    """Excel dosyasından rapor verilerini okur."""

    def __init__(self, excel_path: str):
        """
        Args:
            excel_path: Excel dosyasının yolu
        """
        self.excel_path = excel_path
        self.workbook = None
        self.data = {}

    def load(self) -> bool:
        """Excel dosyasını yükler."""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Dosya bulunamadi: {self.excel_path}")

        self.workbook = load_workbook(self.excel_path)
        return True

    def get_sheet_names(self) -> List[str]:
        """Tüm sayfa isimlerini döndürür."""
        if not self.workbook:
            self.load()
        return self.workbook.sheetnames

    def read_firma_bilgileri(self) -> Dict[str, Any]:
        """FirmaBilgileri sayfasını okur."""
        if not self.workbook:
            self.load()

        if 'FirmaBilgileri' not in self.workbook.sheetnames:
            return {}

        sheet = self.workbook['FirmaBilgileri']
        data = {}

        # A sütunu: alan adları, B sütunu: değerler
        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            if key:
                data[key.strip()] = value if value else ""

        return data

    def read_ana_dagitim_pano(self) -> Dict[str, Any]:
        """AnaDagitimPano sayfasını okur (2.1 DETAY BİLGİLER)."""
        if not self.workbook:
            self.load()

        if 'AnaDagitimPano' not in self.workbook.sheetnames:
            return {}

        sheet = self.workbook['AnaDagitimPano']
        data = {}

        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            if key:
                data[key.strip()] = value if value else ""

        return data

    def read_cihaz_bilgileri(self) -> Dict[str, Any]:
        """CihazBilgileri sayfasını okur."""
        if not self.workbook:
            self.load()

        if 'CihazBilgileri' not in self.workbook.sheetnames:
            return {}

        sheet = self.workbook['CihazBilgileri']
        data = {
            'termal_kamera': {},
            'olcum_aleti': {}
        }

        current_section = None
        for row in range(2, sheet.max_row + 1):
            cell_a = sheet.cell(row=row, column=1).value
            cell_b = sheet.cell(row=row, column=2).value

            if cell_a:
                cell_a = str(cell_a).strip()

                if 'TERMAL' in cell_a.upper():
                    current_section = 'termal_kamera'
                elif 'OLCUM' in cell_a.upper() or 'ÖLÇÜM' in cell_a.upper():
                    current_section = 'olcum_aleti'
                elif current_section and not cell_a.startswith("---"):
                    data[current_section][cell_a] = cell_b if cell_b else ""

        return data

    def read_gozle_kontrol(self) -> Dict[str, Any]:
        """GozleKontrol sayfasını okur (Kusur derecesi dahil)."""
        if not self.workbook:
            self.load()

        if 'GozleKontrol' not in self.workbook.sheetnames:
            return {}

        sheet = self.workbook['GozleKontrol']
        data = {
            'pano_adi': None,
            'kontroller': {},
            'kusurlar': []  # Kusurlu maddeler
        }

        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            kusur = sheet.cell(row=row, column=3).value

            if key:
                key = str(key).strip()

                # Başlıkları atla
                if key.startswith("---"):
                    continue

                if 'PANO_ADI' in key.upper() or 'PANO ADI' in key.upper():
                    data['pano_adi'] = value if value else ""
                else:
                    data['kontroller'][key] = value if value else ""

                    # Uygun değilse kusur listesine ekle
                    if value and 'UYGUN DEĞİL' in str(value).upper():
                        kusur_derece = kusur if kusur else "*"
                        data['kusurlar'].append({
                            'madde': key,
                            'derece': kusur_derece
                        })

        return data

    def read_fonksiyon_testleri(self, ana_pano_data: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """FonksiyonTestleri sayfasını okur, Iz'i otomatik hesaplar.

        NOT: RCD miras sistemi kaldırıldı - boş hücreler boş kalır.
        """
        if not self.workbook:
            self.load()

        if 'FonksiyonTestleri' not in self.workbook.sheetnames:
            return []

        sheet = self.workbook['FonksiyonTestleri']

        # Başlıkları oku (1. satır)
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header if header else f"Col{col}")

        # Verileri oku
        data = []

        for row in range(2, sheet.max_row + 1):
            row_data = {}
            has_data = False

            for col, header in enumerate(headers, 1):
                value = sheet.cell(row=row, column=col).value

                # Başlıkları standardize et (DOCX ve GUI ile uyum için)
                if header == 'Faz Kesiti (mm2)': header = 'Faz Kesiti'
                elif header == 'N Kesiti (mm2)': header = 'Notr Kesiti'
                elif header == 'PE Kesiti (mm2)': header = 'Toprak Kesiti'
                elif header == 'Ib (A)': header = 'Ib'
                elif header == 'Iz (A)': header = 'Iz'
                elif header == 'Icu (kA)': header = 'Icu'

                row_data[header] = value if value else ""
                if value:
                    has_data = True

            if has_data:
                # Iz'i otomatik hesapla (Faz Kesiti'nden)
                faz_kesiti = row_data.get('Faz Kesiti', 0)
                if faz_kesiti:
                    iz_value = get_iz_from_kesit(faz_kesiti)
                    row_data['Iz'] = iz_value

                # RCD değerleri olduğu gibi bırak (miras sistemi kaldırıldı)

                data.append(row_data)

        return data

    def read_termal_goruntuler(self) -> List[Dict[str, str]]:
        """TermalGoruntuler sayfasını okur."""
        if not self.workbook:
            self.load()

        if 'TermalGoruntuler' not in self.workbook.sheetnames:
            return []

        sheet = self.workbook['TermalGoruntuler']
        data = []

        for row in range(2, sheet.max_row + 1):
            pano_adi = sheet.cell(row=row, column=1).value
            fluke_dosya = sheet.cell(row=row, column=2).value

            # fluke_dosya varsa ekle (pano_adi opsiyonel)
            if fluke_dosya:
                data.append({
                    'pano_adi': str(pano_adi).strip() if pano_adi else '',
                    'fluke_dosya': str(fluke_dosya).strip()
                })

        return data

    def read_sonuc(self) -> Dict[str, Any]:
        """Sonuc sayfasını okur."""
        if not self.workbook:
            self.load()

        if 'Sonuc' not in self.workbook.sheetnames:
            return {}

        sheet = self.workbook['Sonuc']
        data = {}

        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=2).value
            if key:
                data[key.strip()] = value if value else ""

        return data

    def read_all(self) -> Dict[str, Any]:
        """Tüm sayfaları okur."""
        self.load()

        # Önce ana pano verisini oku (RCD miras için gerekli)
        ana_pano = self.read_ana_dagitim_pano()

        return {
            'firma_bilgileri': self.read_firma_bilgileri(),
            'ana_dagitim_pano': ana_pano,
            'cihaz_bilgileri': self.read_cihaz_bilgileri(),
            'gozle_kontrol': self.read_gozle_kontrol(),
            'fonksiyon_testleri': self.read_fonksiyon_testleri(ana_pano),  # Ana pano verisini geçir
            'termal_goruntuler': self.read_termal_goruntuler(),
            'sonuc': self.read_sonuc()
        }


# Test için
if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        # Relative path for testing or current directory
        excel_path = "ornek_veri.xlsx"

    if os.path.exists(excel_path):
        reader = ExcelReader(excel_path)
        data = reader.read_all()
        print(json.dumps(data, indent=2, ensure_ascii=False, default=str))
    else:
        print(f"Test dosyasi bulunamadi: {excel_path}")
