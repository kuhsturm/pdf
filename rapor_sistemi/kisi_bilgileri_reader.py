"""
Kişi Bilgileri Okuyucu Modülü
Excel dosyasından kişi adına göre cihaz bilgilerini çeker.
Sözleşmeden gelen kontrol eden kişi adına göre otomatik eşleştirme yapar.
"""

from openpyxl import load_workbook
from typing import Dict, Any, Optional, List
import os
from datetime import datetime


def normalize_name(name: str) -> str:
    """İsmi normalize eder (büyük harf, boşluk düzenleme, Türkçe karakter)."""
    if not name:
        return ""
    # Büyük harfe çevir
    name = name.upper().strip()
    # Türkçe karakterleri ASCII'ye çevir (karşılaştırma için)
    tr_map = {
        'İ': 'I', 'Ğ': 'G', 'Ü': 'U', 'Ş': 'S', 'Ö': 'O', 'Ç': 'C',
        'ı': 'I', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c'
    }
    for tr_char, ascii_char in tr_map.items():
        name = name.replace(tr_char, ascii_char)
    # Çoklu boşlukları tek boşluğa indir
    name = ' '.join(name.split())
    return name


def format_date(val) -> str:
    """Tarih değerini DD.MM.YYYY formatına çevirir."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, str):
        # Zaten string ise olduğu gibi döndür
        return val
    return str(val)


class KisiBilgileriReader:
    """Excel dosyasından kişi ve cihaz bilgilerini okur."""

    # Sütun indeksleri (1-indexed)
    NAME_COL = 4  # D sütunu - kişi adı ve alan adları
    VALUE_COL = 5  # E sütunu - değerler

    # Kişi listesi sütunları
    PERSON_NAME_COL = 8  # H sütunu
    PERSON_TC_COL = 9    # I sütunu

    def __init__(self, excel_path: str):
        """
        Args:
            excel_path: kisi_bilgileri.xlsx dosyasının yolu
        """
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        self.persons = {}  # {normalize_name: {row_start, name, tc, cihaz_bilgileri}}

    def load(self) -> bool:
        """Excel dosyasını yükler."""
        if not os.path.exists(self.excel_path):
            print(f"Dosya bulunamadı: {self.excel_path}")
            return False

        try:
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook.active
            self._parse_persons()
            return True
        except Exception as e:
            print(f"Excel okuma hatası: {e}")
            return False

    def _parse_persons(self):
        """Tüm kişileri ve cihaz bilgilerini parse eder.
        Dinamik olarak 'TERMAL KAMERA' başlıklarını arar.
        """
        ws = self.worksheet

        # Önce sağ taraftaki kişi listesini oku (H-I sütunları)
        person_list = []
        # Sağ taraftaki kişi listesi için dinamik arama yapalım veya sabit 2-12 varsayalım
        # Daha güvenli olması için boş olmayan satırları okuyalım
        for row_idx in range(2, 50): # Makul bir sınır
            name_cell = ws.cell(row=row_idx, column=self.PERSON_NAME_COL).value
            tc_cell = ws.cell(row=row_idx, column=self.PERSON_TC_COL).value
            if name_cell:
                person_list.append({
                    'name': str(name_cell).strip(),
                    'tc': str(tc_cell) if tc_cell else ""
                })

        # Sol taraftaki cihaz bilgisi bloklarını parse et
        # Sabit blok boyutu yerine dinamik arama yapıyoruz
        max_row = ws.max_row

        # 'TERMAL KAMERA' başlığını içeren satırları bul
        block_starts = []
        for row_idx in range(1, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=self.NAME_COL).value
            if cell_val and "TERMAL KAMERA" in str(cell_val).upper():
                block_starts.append(row_idx)

        for i, start_row in enumerate(block_starts):
            # Kişi adı genellikle başlığın bir alt satırındadır
            person_row = start_row + 1
            person_name_val = ws.cell(row=person_row, column=self.NAME_COL).value

            if person_name_val:
                person_name = str(person_name_val).strip()
                normalized = normalize_name(person_name)

                # Bir sonraki bloğun başlangıcına kadar veya dosya sonuna kadar oku
                end_row = block_starts[i+1] if i + 1 < len(block_starts) else max_row + 1

                # Cihaz bilgilerini oku (blok içindeki veriler)
                cihaz_data = self._read_cihaz_block_dynamic(start_row, end_row)

                # TC numarasını kişi listesinden bul
                tc_no = ""
                for p in person_list:
                    if normalize_name(p['name']) == normalized:
                        tc_no = p['tc']
                        break

                self.persons[normalized] = {
                    'row_start': start_row,
                    'name': person_name,
                    'tc': tc_no,
                    'cihaz_bilgileri': cihaz_data
                }

    def _read_cihaz_block_dynamic(self, start_row: int, end_row: int) -> Dict[str, Any]:
        """Bir kişinin cihaz bilgileri bloğunu dinamik olarak okur."""
        ws = self.worksheet
        data = {}

        # Anahtar kelimeler ve eşleşen alan adları
        # (Excel hücresindeki metin parçası, kod içindeki alan adı)
        keywords = [
            ('TERMAL CİHAZ', 'termal_cihaz_adi'),
            ('TERMAL KALİBRASYON TARİHİ', 'termal_kalibrasyon_tarihi'),
            ('TERMAL GEÇERLİLİK', 'termal_kalibrasyon_gecerlilik'),
            ('TERMAL SERİ', 'termal_seri_numarasi'),
            ('TERMAL KALİBRASYON NO', 'termal_kalibrasyon_no'),
            ('ÖLÇÜM CİHAZ', 'olcum_cihaz_adi'),
            ('ÖLÇÜM KALİBRASYON TARİHİ', 'olcum_kalibrasyon_tarihi'),
            ('ÖLÇÜM GEÇERLİLİK', 'olcum_kalibrasyon_gecerlilik'),
            ('ÖLÇÜM SERİ', 'olcum_seri_numarasi'),
            ('ÖLÇÜM KALİBRASYON NO', 'olcum_kalibrasyon_no'),
        ]

        for row in range(start_row, end_row):
            key_cell = ws.cell(row=row, column=self.NAME_COL).value
            val_cell = ws.cell(row=row, column=self.VALUE_COL).value

            if key_cell:
                key_text = str(key_cell).upper().replace('İ', 'I').replace('Ç', 'C').replace('Ş', 'S').replace('Ğ', 'G').replace('Ü', 'U').replace('Ö', 'O')

                for keyword, field_name in keywords:
                    normalized_keyword = keyword.replace('İ', 'I').replace('Ç', 'C').replace('Ş', 'S').replace('Ğ', 'G').replace('Ü', 'U').replace('Ö', 'O')

                    if normalized_keyword in key_text:
                        # Değer var mı?
                        if val_cell and str(val_cell).strip() != str(key_cell).strip():
                            if 'tarihi' in field_name or 'gecerlilik' in field_name:
                                data[field_name] = format_date(val_cell)
                            else:
                                data[field_name] = str(val_cell)
        return data

    def get_person_list(self) -> List[str]:
        """Tüm kişi adlarını döndürür."""
        return [p['name'] for p in self.persons.values()]

    def get_person_by_name(self, name: str) -> Optional[Dict[str, Any]]:
        """
        İsme göre kişi bilgilerini döndürür.
        Fuzzy matching yapar (Türkçe karakter ve büyük/küçük harf duyarsız).
        """
        normalized = normalize_name(name)

        # Tam eşleşme
        if normalized in self.persons:
            return self.persons[normalized]

        # Kısmi eşleşme (isim içerme kontrolü)
        for key, person in self.persons.items():
            if normalized in key or key in normalized:
                return person

        return None

    def get_cihaz_bilgileri(self, name: str) -> Dict[str, Any]:
        """
        İsme göre cihaz bilgilerini döndürür.
        Sözleşmeden gelen kontrol_eden_adsoyad ile eşleştirmek için kullanılır.
        """
        person = self.get_person_by_name(name)
        if person:
            return person.get('cihaz_bilgileri', {})
        return {}

    def get_tc_no(self, name: str) -> str:
        """İsme göre TC numarasını döndürür."""
        person = self.get_person_by_name(name)
        if person:
            return person.get('tc', '')
        return ''


def get_cihaz_from_sozlesme(kisi_excel_path: str, sozlesme_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Sözleşme verisindeki kontrol eden kişiye göre cihaz bilgilerini döndürür.

    Args:
        kisi_excel_path: kisi_bilgileri.xlsx dosya yolu
        sozlesme_data: parse_sozlesme_pdf() fonksiyonundan dönen veri

    Returns:
        Cihaz bilgilerini içeren dict
    """
    kontrol_eden = sozlesme_data.get('kontrol_eden_adsoyad', '')
    if not kontrol_eden:
        return {}

    reader = KisiBilgileriReader(kisi_excel_path)
    if not reader.load():
        return {}

    cihaz = reader.get_cihaz_bilgileri(kontrol_eden)

    # TC numarasını da ekle
    tc = reader.get_tc_no(kontrol_eden)
    if tc:
        cihaz['kontrol_eden_tc'] = tc

    return cihaz


# Test için
if __name__ == "__main__":
    import json

    excel_path = "kisi_bilgileri.xlsx"

    reader = KisiBilgileriReader(excel_path)
    if reader.load():
        print("=== BULUNAN KİŞİLER ===")
        for name in reader.get_person_list():
            print(f"  - {name}")

        print("\n=== ERGİN ÇOŞKUN CİHAZ BİLGİLERİ ===")
        cihaz = reader.get_cihaz_bilgileri("ERGİN ÇOŞKUN")
        print(json.dumps(cihaz, indent=2, ensure_ascii=False))

        print("\n=== AHMET IŞIK CİHAZ BİLGİLERİ ===")
        cihaz = reader.get_cihaz_bilgileri("AHMET IŞIK")
        print(json.dumps(cihaz, indent=2, ensure_ascii=False))
